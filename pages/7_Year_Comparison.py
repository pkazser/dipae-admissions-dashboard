import streamlit as st
import plotly.express as px
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from modules.database_manager import load_admissions_with_departments


st.set_page_config(
    page_title="Σύγκριση Ετών | ΔΙΠΑΕ",
    page_icon="📅",
    layout="wide"
)


st.title("📅 Σύγκριση Ετών Εισακτέων ΔΙ.ΠΑ.Ε.")

st.markdown("""
Σε αυτή τη σελίδα συγκρίνουμε δύο έτη εισαγωγής και εντοπίζουμε μεταβολές
σε θέσεις, επιτυχόντες, κάλυψη και βάσεις εισαγωγής.
""")

st.divider()


def apply_category_filter(df, selected_category):
    """
    Εφαρμόζει κοινή λογική φίλτρου κατηγορίας.
    """

    if selected_category == "Όλες οι κατηγορίες":
        return df.copy()

    if selected_category == "Βασικές κατηγορίες χωρίς 10%":
        return df[
            ~df["exam_category"].astype(str).str.contains("10%", na=False)
        ].copy()

    if selected_category == "Μόνο 10%":
        return df[
            df["exam_category"].astype(str).str.contains("10%", na=False)
        ].copy()

    return df[df["exam_category"] == selected_category].copy()


def summarize_by_department(df):
    """
    Σύνοψη ανά Τμήμα για ένα έτος.

    Για θέσεις και επιτυχόντες γίνεται άθροισμα.
    Για μόρια πρώτου και βάση τελευταίου παίρνουμε μέσο όρο,
    άρα είναι ενδεικτικό όταν υπάρχουν πολλές κατηγορίες.
    """

    summary = (
        df
        .groupby(
            [
                "department_code",
                "department_name_clean",
                "school",
                "city",
            ],
            as_index=False
        )
        .agg(
            final_positions=("final_positions", "sum"),
            admitted=("admitted", "sum"),
            first_score=("first_score", "mean"),
            base_score=("base_score", "mean"),
        )
    )

    summary["coverage"] = (
        summary["admitted"] / summary["final_positions"] * 100
    )

    return summary


def summarize_by_field(df, field):
    """
    Σύνοψη ανά Σχολή ή Πόλη.
    """

    summary = (
        df
        .groupby(field, as_index=False)
        .agg(
            departments=("department_name_clean", "nunique"),
            final_positions=("final_positions", "sum"),
            admitted=("admitted", "sum"),
            first_score=("first_score", "mean"),
            base_score=("base_score", "mean"),
        )
    )

    summary["coverage"] = (
        summary["admitted"] / summary["final_positions"] * 100
    )

    return summary


def compare_summaries(df_old, df_new, key_columns, suffix_old, suffix_new):
    """
    Ενώνει δύο summaries και υπολογίζει διαφορές.
    """

    merged = df_old.merge(
        df_new,
        on=key_columns,
        how="inner",
        suffixes=(f"_{suffix_old}", f"_{suffix_new}")
    )

    merged["diff_final_positions"] = (
        merged[f"final_positions_{suffix_new}"]
        - merged[f"final_positions_{suffix_old}"]
    )

    merged["diff_admitted"] = (
        merged[f"admitted_{suffix_new}"]
        - merged[f"admitted_{suffix_old}"]
    )

    merged["diff_coverage"] = (
        merged[f"coverage_{suffix_new}"]
        - merged[f"coverage_{suffix_old}"]
    )

    merged["diff_base_score"] = (
        merged[f"base_score_{suffix_new}"]
        - merged[f"base_score_{suffix_old}"]
    )

    merged["diff_first_score"] = (
        merged[f"first_score_{suffix_new}"]
        - merged[f"first_score_{suffix_old}"]
    )

    return merged


def highlight_differences(val):
    """
    Χρωματίζει τις μεταβολές:
    θετικές πράσινο, αρνητικές κόκκινο, μηδενικές ουδέτερο.
    """

    try:
        value = float(val)
    except Exception:
        return ""

    if value > 0:
        return "background-color: #d4edda; color: #155724; font-weight: bold;"
    elif value < 0:
        return "background-color: #f8d7da; color: #721c24; font-weight: bold;"
    else:
        return "background-color: #f1f3f5; color: #495057;"


def style_difference_table(df, diff_columns):
    """
    Εφαρμόζει χρωματισμό στις στήλες διαφορών.
    Χρησιμοποιεί .map για νεότερες εκδόσεις pandas.
    Αν υπάρχει παλιότερη έκδοση pandas, γυρίζει σε applymap.
    """

    try:
        return df.style.map(
            highlight_differences,
            subset=diff_columns
        )
    except AttributeError:
        return df.style.applymap(
            highlight_differences,
            subset=diff_columns
        )


def create_excel_export(
    dept_display,
    school_display,
    city_display,
    year_old,
    year_new,
    selected_category
):
    """
    Δημιουργεί Excel αρχείο σε μνήμη με τα αποτελέσματα σύγκρισης.
    Περιλαμβάνει φύλλα:
    - Τμήματα
    - Σχολές
    - Πόλεις
    """

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dept_display.to_excel(
            writer,
            index=False,
            sheet_name="Τμήματα",
            startrow=4
        )

        school_display.to_excel(
            writer,
            index=False,
            sheet_name="Σχολές",
            startrow=4
        )

        city_display.to_excel(
            writer,
            index=False,
            sheet_name="Πόλεις",
            startrow=4
        )

        workbook = writer.book

        for sheet_name in ["Τμήματα", "Σχολές", "Πόλεις"]:
            worksheet = workbook[sheet_name]

            worksheet["A1"] = "Σύγκριση Ετών Εισακτέων ΔΙ.ΠΑ.Ε."
            worksheet["A2"] = f"Έτη: {year_old} → {year_new}"
            worksheet["A3"] = f"Κατηγορία: {selected_category}"

            worksheet["A1"].font = Font(bold=True, size=14)
            worksheet["A2"].font = Font(bold=True)
            worksheet["A3"].font = Font(bold=True)

            header_row = 5

            header_fill = PatternFill(
                start_color="D9EAF7",
                end_color="D9EAF7",
                fill_type="solid"
            )

            for cell in worksheet[header_row]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Χρωματισμός στηλών διαφορών στο Excel
            green_fill = PatternFill(
                start_color="D4EDDA",
                end_color="D4EDDA",
                fill_type="solid"
            )

            red_fill = PatternFill(
                start_color="F8D7DA",
                end_color="F8D7DA",
                fill_type="solid"
            )

            neutral_fill = PatternFill(
                start_color="F1F3F5",
                end_color="F1F3F5",
                fill_type="solid"
            )

            diff_headers = [
                "Διαφορά Θέσεων",
                "Διαφορά Επιτυχόντων",
                "Διαφορά Κάλυψης",
                "Διαφορά Βάσης",
            ]

            header_map = {}

            for cell in worksheet[header_row]:
                header_map[cell.value] = cell.column

            for diff_header in diff_headers:
                if diff_header in header_map:
                    col_idx = header_map[diff_header]

                    for row in range(header_row + 1, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_idx)

                        try:
                            value = float(cell.value)
                        except Exception:
                            continue

                        if value > 0:
                            cell.fill = green_fill
                        elif value < 0:
                            cell.fill = red_fill
                        else:
                            cell.fill = neutral_fill

            # Αυτόματο πλάτος στηλών
            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    try:
                        value = "" if cell.value is None else str(cell.value)
                        if len(value) > max_length:
                            max_length = len(value)
                    except Exception:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            worksheet.freeze_panes = "A6"

    output.seek(0)
    return output


try:
    df = load_admissions_with_departments()

    if df.empty:
        st.warning("Δεν υπάρχουν ακόμη δεδομένα εισακτέων στη βάση.")
        st.stop()

    years = sorted(df["year"].dropna().unique().tolist())

    if len(years) < 2:
        st.warning("Χρειάζονται τουλάχιστον δύο έτη στη βάση για σύγκριση.")
        st.stop()

    col_year1, col_year2, col_category = st.columns(3)

    with col_year1:
        year_old = st.selectbox(
            "Παλαιότερο έτος",
            years,
            index=0
        )

    with col_year2:
        year_new = st.selectbox(
            "Νεότερο έτος",
            years,
            index=len(years) - 1
        )

    df_between_years = df[df["year"].isin([year_old, year_new])].copy()

    category_options = [
        "ΓΕΛ Ημερήσια",
        "Βασικές κατηγορίες χωρίς 10%",
        "Όλες οι κατηγορίες",
        "Μόνο 10%",
    ] + sorted(df_between_years["exam_category"].dropna().unique().tolist())

    category_options = list(dict.fromkeys(category_options))

    with col_category:
        selected_category = st.selectbox(
            "Κατηγορία σύγκρισης",
            category_options
        )

    if year_old == year_new:
        st.warning("Επίλεξε δύο διαφορετικά έτη.")
        st.stop()

    df_old = df[df["year"] == year_old].copy()
    df_new = df[df["year"] == year_new].copy()

    df_old = apply_category_filter(df_old, selected_category)
    df_new = apply_category_filter(df_new, selected_category)

    if df_old.empty or df_new.empty:
        st.warning("Δεν υπάρχουν δεδομένα και για τα δύο έτη με τα επιλεγμένα φίλτρα.")
        st.stop()

    st.subheader(f"Σύγκριση {year_old} → {year_new} — {selected_category}")

    # ---------------------------------------------------------
    # Κεντρικά KPIs
    # ---------------------------------------------------------

    old_positions = int(df_old["final_positions"].sum())
    new_positions = int(df_new["final_positions"].sum())

    old_admitted = int(df_old["admitted"].sum())
    new_admitted = int(df_new["admitted"].sum())

    old_coverage = old_admitted / old_positions * 100 if old_positions > 0 else 0
    new_coverage = new_admitted / new_positions * 100 if new_positions > 0 else 0

    old_avg_base = df_old["base_score"].mean()
    new_avg_base = df_new["base_score"].mean()

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric(
            "Τελικές θέσεις",
            new_positions,
            delta=new_positions - old_positions
        )

    with col2:
        st.metric(
            "Επιτυχόντες",
            new_admitted,
            delta=new_admitted - old_admitted
        )

    with col3:
        st.metric(
            "Συνολική κάλυψη",
            f"{new_coverage:.1f}%",
            delta=f"{new_coverage - old_coverage:.1f}%"
        )

    with col4:
        st.metric(
            "Μέση βάση",
            f"{new_avg_base:.0f}",
            delta=f"{new_avg_base - old_avg_base:.0f}"
        )

    st.caption(
        "Οι μεταβολές βάσεων είναι ενδεικτικές όταν η επιλογή περιλαμβάνει πολλές κατηγορίες. "
        "Για καθαρή σύγκριση βάσεων προτίμησε συγκεκριμένη κατηγορία, π.χ. ΓΕΛ Ημερήσια."
    )

    st.divider()

    # ---------------------------------------------------------
    # Σύγκριση ανά Τμήμα
    # ---------------------------------------------------------

    st.subheader("Σύγκριση ανά Τμήμα")

    old_dept = summarize_by_department(df_old)
    new_dept = summarize_by_department(df_new)

    dept_comparison = compare_summaries(
        old_dept,
        new_dept,
        key_columns=[
            "department_code",
            "department_name_clean",
            "school",
            "city",
        ],
        suffix_old=str(year_old),
        suffix_new=str(year_new)
    )

    dept_display = dept_comparison[
        [
            "department_name_clean",
            "school",
            "city",
            f"final_positions_{year_old}",
            f"final_positions_{year_new}",
            "diff_final_positions",
            f"admitted_{year_old}",
            f"admitted_{year_new}",
            "diff_admitted",
            f"coverage_{year_old}",
            f"coverage_{year_new}",
            "diff_coverage",
            f"base_score_{year_old}",
            f"base_score_{year_new}",
            "diff_base_score",
        ]
    ].copy()

    dept_display = dept_display.rename(
        columns={
            "department_name_clean": "Τμήμα",
            "school": "Σχολή",
            "city": "Πόλη",
            f"final_positions_{year_old}": f"Θέσεις {year_old}",
            f"final_positions_{year_new}": f"Θέσεις {year_new}",
            "diff_final_positions": "Διαφορά Θέσεων",
            f"admitted_{year_old}": f"Επιτυχόντες {year_old}",
            f"admitted_{year_new}": f"Επιτυχόντες {year_new}",
            "diff_admitted": "Διαφορά Επιτυχόντων",
            f"coverage_{year_old}": f"Κάλυψη {year_old} %",
            f"coverage_{year_new}": f"Κάλυψη {year_new} %",
            "diff_coverage": "Διαφορά Κάλυψης",
            f"base_score_{year_old}": f"Βάση {year_old}",
            f"base_score_{year_new}": f"Βάση {year_new}",
            "diff_base_score": "Διαφορά Βάσης",
        }
    )

    numeric_round_columns_dept = [
        f"Κάλυψη {year_old} %",
        f"Κάλυψη {year_new} %",
        "Διαφορά Κάλυψης",
        f"Βάση {year_old}",
        f"Βάση {year_new}",
        "Διαφορά Βάσης",
    ]

    for col in numeric_round_columns_dept:
        dept_display[col] = dept_display[col].round(2)

    diff_columns_dept = [
        "Διαφορά Θέσεων",
        "Διαφορά Επιτυχόντων",
        "Διαφορά Κάλυψης",
        "Διαφορά Βάσης",
    ]

    st.dataframe(
        style_difference_table(dept_display, diff_columns_dept),
        use_container_width=True,
        hide_index=True
    )

    st.divider()

    # ---------------------------------------------------------
    # Γραφήματα μεταβολών ανά Τμήμα
    # ---------------------------------------------------------

    st.subheader("Γραφήματα μεταβολών ανά Τμήμα")

    col_chart_a, col_chart_b = st.columns(2)

    with col_chart_a:
        top_base_increase = dept_display.sort_values(
            "Διαφορά Βάσης",
            ascending=False
        ).head(10)

        fig_base_up = px.bar(
            top_base_increase,
            x="Τμήμα",
            y="Διαφορά Βάσης",
            text="Διαφορά Βάσης",
            hover_data=["Σχολή", "Πόλη"],
            title="Top-10 αυξήσεις βάσης"
        )

        fig_base_up.update_traces(textposition="outside")

        fig_base_up.update_layout(
            xaxis_title="Τμήμα",
            yaxis_title="Διαφορά βάσης",
            uniformtext_minsize=8,
            uniformtext_mode="hide"
        )

        st.plotly_chart(
            fig_base_up,
            use_container_width=True
        )

    with col_chart_b:
        top_base_decrease = dept_display.sort_values(
            "Διαφορά Βάσης",
            ascending=True
        ).head(10)

        fig_base_down = px.bar(
            top_base_decrease,
            x="Τμήμα",
            y="Διαφορά Βάσης",
            text="Διαφορά Βάσης",
            hover_data=["Σχολή", "Πόλη"],
            title="Top-10 μειώσεις βάσης"
        )

        fig_base_down.update_traces(textposition="outside")

        fig_base_down.update_layout(
            xaxis_title="Τμήμα",
            yaxis_title="Διαφορά βάσης",
            uniformtext_minsize=8,
            uniformtext_mode="hide"
        )

        st.plotly_chart(
            fig_base_down,
            use_container_width=True
        )

    col_chart_c, col_chart_d = st.columns(2)

    with col_chart_c:
        top_admitted_increase = dept_display.sort_values(
            "Διαφορά Επιτυχόντων",
            ascending=False
        ).head(10)

        fig_admitted_up = px.bar(
            top_admitted_increase,
            x="Τμήμα",
            y="Διαφορά Επιτυχόντων",
            text="Διαφορά Επιτυχόντων",
            hover_data=["Σχολή", "Πόλη"],
            title="Top-10 αυξήσεις επιτυχόντων"
        )

        fig_admitted_up.update_traces(textposition="outside")

        fig_admitted_up.update_layout(
            xaxis_title="Τμήμα",
            yaxis_title="Διαφορά επιτυχόντων",
            uniformtext_minsize=8,
            uniformtext_mode="hide"
        )

        st.plotly_chart(
            fig_admitted_up,
            use_container_width=True
        )

    with col_chart_d:
        top_admitted_decrease = dept_display.sort_values(
            "Διαφορά Επιτυχόντων",
            ascending=True
        ).head(10)

        fig_admitted_down = px.bar(
            top_admitted_decrease,
            x="Τμήμα",
            y="Διαφορά Επιτυχόντων",
            text="Διαφορά Επιτυχόντων",
            hover_data=["Σχολή", "Πόλη"],
            title="Top-10 μειώσεις επιτυχόντων"
        )

        fig_admitted_down.update_traces(textposition="outside")

        fig_admitted_down.update_layout(
            xaxis_title="Τμήμα",
            yaxis_title="Διαφορά επιτυχόντων",
            uniformtext_minsize=8,
            uniformtext_mode="hide"
        )

        st.plotly_chart(
            fig_admitted_down,
            use_container_width=True
        )

    st.divider()

    # ---------------------------------------------------------
    # Σύγκριση ανά Σχολή και Πόλη
    # ---------------------------------------------------------

    st.subheader("Σύγκριση ανά Σχολή και Πόλη")

    tab_school, tab_city = st.tabs(["Ανά Σχολή", "Ανά Πόλη"])

    with tab_school:
        old_school = summarize_by_field(df_old, "school")
        new_school = summarize_by_field(df_new, "school")

        school_comparison = compare_summaries(
            old_school,
            new_school,
            key_columns=["school"],
            suffix_old=str(year_old),
            suffix_new=str(year_new)
        )

        school_display = school_comparison[
            [
                "school",
                f"final_positions_{year_old}",
                f"final_positions_{year_new}",
                "diff_final_positions",
                f"admitted_{year_old}",
                f"admitted_{year_new}",
                "diff_admitted",
                f"coverage_{year_old}",
                f"coverage_{year_new}",
                "diff_coverage",
                f"base_score_{year_old}",
                f"base_score_{year_new}",
                "diff_base_score",
            ]
        ].copy()

        school_display = school_display.rename(
            columns={
                "school": "Σχολή",
                f"final_positions_{year_old}": f"Θέσεις {year_old}",
                f"final_positions_{year_new}": f"Θέσεις {year_new}",
                "diff_final_positions": "Διαφορά Θέσεων",
                f"admitted_{year_old}": f"Επιτυχόντες {year_old}",
                f"admitted_{year_new}": f"Επιτυχόντες {year_new}",
                "diff_admitted": "Διαφορά Επιτυχόντων",
                f"coverage_{year_old}": f"Κάλυψη {year_old} %",
                f"coverage_{year_new}": f"Κάλυψη {year_new} %",
                "diff_coverage": "Διαφορά Κάλυψης",
                f"base_score_{year_old}": f"Βάση {year_old}",
                f"base_score_{year_new}": f"Βάση {year_new}",
                "diff_base_score": "Διαφορά Βάσης",
            }
        )

        numeric_round_columns_school = [
            f"Κάλυψη {year_old} %",
            f"Κάλυψη {year_new} %",
            "Διαφορά Κάλυψης",
            f"Βάση {year_old}",
            f"Βάση {year_new}",
            "Διαφορά Βάσης",
        ]

        for col in numeric_round_columns_school:
            school_display[col] = school_display[col].round(2)

        diff_columns_school = [
            "Διαφορά Θέσεων",
            "Διαφορά Επιτυχόντων",
            "Διαφορά Κάλυψης",
            "Διαφορά Βάσης",
        ]

        st.dataframe(
            style_difference_table(school_display, diff_columns_school),
            use_container_width=True,
            hide_index=True
        )

        fig_school = px.bar(
            school_display,
            x="Σχολή",
            y="Διαφορά Επιτυχόντων",
            text="Διαφορά Επιτυχόντων",
            title="Μεταβολή επιτυχόντων ανά Σχολή"
        )

        fig_school.update_traces(textposition="outside")

        fig_school.update_layout(
            xaxis_title="Σχολή",
            yaxis_title="Διαφορά επιτυχόντων",
            uniformtext_minsize=8,
            uniformtext_mode="hide"
        )

        st.plotly_chart(
            fig_school,
            use_container_width=True
        )

    with tab_city:
        old_city = summarize_by_field(df_old, "city")
        new_city = summarize_by_field(df_new, "city")

        city_comparison = compare_summaries(
            old_city,
            new_city,
            key_columns=["city"],
            suffix_old=str(year_old),
            suffix_new=str(year_new)
        )

        city_display = city_comparison[
            [
                "city",
                f"final_positions_{year_old}",
                f"final_positions_{year_new}",
                "diff_final_positions",
                f"admitted_{year_old}",
                f"admitted_{year_new}",
                "diff_admitted",
                f"coverage_{year_old}",
                f"coverage_{year_new}",
                "diff_coverage",
                f"base_score_{year_old}",
                f"base_score_{year_new}",
                "diff_base_score",
            ]
        ].copy()

        city_display = city_display.rename(
            columns={
                "city": "Πόλη",
                f"final_positions_{year_old}": f"Θέσεις {year_old}",
                f"final_positions_{year_new}": f"Θέσεις {year_new}",
                "diff_final_positions": "Διαφορά Θέσεων",
                f"admitted_{year_old}": f"Επιτυχόντες {year_old}",
                f"admitted_{year_new}": f"Επιτυχόντες {year_new}",
                "diff_admitted": "Διαφορά Επιτυχόντων",
                f"coverage_{year_old}": f"Κάλυψη {year_old} %",
                f"coverage_{year_new}": f"Κάλυψη {year_new} %",
                "diff_coverage": "Διαφορά Κάλυψης",
                f"base_score_{year_old}": f"Βάση {year_old}",
                f"base_score_{year_new}": f"Βάση {year_new}",
                "diff_base_score": "Διαφορά Βάσης",
            }
        )

        numeric_round_columns_city = [
            f"Κάλυψη {year_old} %",
            f"Κάλυψη {year_new} %",
            "Διαφορά Κάλυψης",
            f"Βάση {year_old}",
            f"Βάση {year_new}",
            "Διαφορά Βάσης",
        ]

        for col in numeric_round_columns_city:
            city_display[col] = city_display[col].round(2)

        diff_columns_city = [
            "Διαφορά Θέσεων",
            "Διαφορά Επιτυχόντων",
            "Διαφορά Κάλυψης",
            "Διαφορά Βάσης",
        ]

        st.dataframe(
            style_difference_table(city_display, diff_columns_city),
            use_container_width=True,
            hide_index=True
        )

        fig_city = px.bar(
            city_display,
            x="Πόλη",
            y="Διαφορά Επιτυχόντων",
            text="Διαφορά Επιτυχόντων",
            title="Μεταβολή επιτυχόντων ανά Πόλη"
        )

        fig_city.update_traces(textposition="outside")

        fig_city.update_layout(
            xaxis_title="Πόλη",
            yaxis_title="Διαφορά επιτυχόντων",
            uniformtext_minsize=8,
            uniformtext_mode="hide"
        )

        st.plotly_chart(
            fig_city,
            use_container_width=True
        )

    st.divider()

    # ---------------------------------------------------------
    # Export Excel
    # ---------------------------------------------------------

    st.subheader("Εξαγωγή αποτελεσμάτων")

    excel_file = create_excel_export(
        dept_display=dept_display,
        school_display=school_display,
        city_display=city_display,
        year_old=year_old,
        year_new=year_new,
        selected_category=selected_category
    )

    st.download_button(
        label="📥 Κατέβασμα Excel σύγκρισης",
        data=excel_file,
        file_name=f"dipae_year_comparison_{year_old}_{year_new}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

except Exception as e:
    st.error("Υπήρξε πρόβλημα κατά τη σύγκριση ετών.")
    st.exception(e)